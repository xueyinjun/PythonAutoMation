# 导入Workbook类开始工作
# from openpyxl import Workbook
# wb = Workbook()

#使用active属性获取它
# ws = wb.active

#Workbook.create_sheet()方法创建新工作表
# ws1 = wb.create_sheet("Mysheet")
# ws2 = wb.create_sheet("Mysheet", 0)

#Worksheet.title属性修改顺序编号名称
# ws.title = "New Title"

# ws.sheet_properties.tabColor = "1072BA"

#工作表提供了名称，就可以将其作为工作薄的一个键
# ws3 = wb["New Title"]

#查看所有工作表的名称
# print(wb.sheetnames)

#遍历工作表
# for sheet in wb:
#     print(sheet.title)

#从单个工作簿中创建工作表的方法
# source = wb.active
# target = wb.copy_worksheet(source)

#使用openpyxl.load_workbook() 打开现在工作簿

# from openpyxl import load_workbook
# wb4 = load_workbook("test.xlsx")
# print(wb4.sheetnames)

# ws4 = wb4.active

#单元格可以直接作为工作表的键访问，如果单元格不存在，则创建一个单元格，值可以直接分配

#Worksheet.cell()方法
# d = ws4.cell(row=4, column=2, value=4)
# c = ws4.cell(row=5, column=2, value=5)
#访问多个单元格范围 使用切片的方式
# cell_range = ws4['A1':'C2']

#行或列的范围可以类似的获得
# colC = ws4['C']
# col_range = ws4['C:D']
# row10 = ws4[10]
# row_range = ws4[5:10]

#使用Worksheet.iter_rows()方法 访问多个单元格
# for row in ws4.iter_rows(min_row=1, max_col=3, max_row=2):
#     for cell in row:
#         print(cell)

# for col in ws4.iter_cols(min_row=1, max_col=3, max_row=2):
#     for cell in col:
#         print(cell)

#需要遍历文件的所有行或者列，Worksheet.rows属性,Worksheet.columns
# ws4 = wb.active
# ws4['C9'] = 'hello world'
# print(tuple(ws4.rows))
# print(tuple(ws4.columns))

#仅限值 Worksheet.values属性 这回迭代所有工作表中的行,但只返回单元格值

# for row in ws4.values:
#     for value in row:
#         print(value)

# Worksheet.iter_rows()并Worksheet.iter_cols()可以采用values_only参数,只返回单元格到的值
# for row in ws4.iter_rows(min_row=1, max_col=3, max_row=2, values_only=True):
#     print(row)

#我们有了cell 可以为它分配一个值
# d.value = 'hello,world'
# print(d.value)
# c.value = 3.14
# print(c.value)

#保存到文件
# wb1= Workbook()
# wb1.save('test.xlsx')

#将文件保存到流中, NameTemporaryFile()
# from tempfile import NamedTemporaryFile
# from openpyxl import Workbook
# from openpyxl import load_workbook
# wb = Workbook()
# with NamedTemporaryFile() as tmp:
#     wb.save(tmp.name)
#     tmp.seek(0)
#     stream = tmp.read()

#可以指定属性 template= True 以将工作簿另存为模板
# wb = load_workbook('document.xlsx')
# wb.template =True
# wb.save('document.xlsx',as_template=False)

#从文件加载，使用load_workbook打开现有工作簿
# from openpyxl import load_workbook
# wb2 = load_workbook('test.xlsx')
# print(wb2.sheetnames)

# from openpyxl import load_workbook
# wb  = load_workbook('test.xlsx')

# ws  = wb.active

# for row in ws.iter_rows(min_row=1,max_col=3,max_row=2):
#     for cell in row:
#         print(cell)
# for col in ws.iter_cols(min_rowj=1,max_col=3,max_row=2):
#     for cell in col:
#         print(cell)

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb = Workbook()

dest_filename = 'empty_book.xlsx'

ws1 = wb.active
ws1.title = "range names"

for row in range(1, 40):
    ws1.append.range((600))

ws2 = wb.create_sheet(title="Pi")
ws2['F5'] = 3.14

ws3 = wb.create_chartsheet(title="Data")
for row in range(10, 20):
    for col in range(27, 54):
        ws3.cell(
            column=col, row=row, value="{0}".format(get_column_letter(col)))
print(ws3['AA10'].value)
wb.save(filename=dest_filename)
