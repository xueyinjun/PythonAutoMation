#utf-8

import openpyxl, smtplib, sys

#Open the spreadsheet and get the latest dues status

wb = openpyxl.load_workbook(
    '/Users/VON/Documents/Python/automate_online-materials/duesRecords.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
lastCol = sheet.max_column
latestMonth = sheet.cell(row=1, column=lastCol).value

#Check each member's payment status.
unpaidMemebers = {}
for r in range(2, sheet.maxrow + 1):
    payment = sheet.cell(row=r, column=lastCol).value
    if payment != 'paid':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaidMemebers[name] = email

#log in to email account

smtpObj = smtpObj.SMTP('smtp.gmail.com', 587)
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login('my_email_address@gmail.com', sys.argv[1])

#Send  out reminder emails
for name, email in unpaidMemebers.items():
    body = "Subject:%s dues unpid\nDear %s, \nRecords show that you have not paid dues for %s.Please make this payment as soon as possible.Thank you!'" % (
        latestMonth, name, latestMonth)
    print('Send email to %s...' % email)
    sendmailStatus = smtpObj.sendmail('my_email_address@gamail.com', email,
                                      body)

    if sendmailStatus != {}:
        print('There was a problem sending email to %s: %s ' %
              (email, sendmailStatus))

smtpObj.quit()

